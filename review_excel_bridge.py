"""Convert minimal demo review tasks between JSONL and Excel."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook, load_workbook

from human_review.io_utils import load_jsonl, write_jsonl


HEADERS = [
    "item_id",
    "target_error_type",
    "scenario_type",
    "plan_id",
    "query",
    "context_preview",
    "reference_answer",
    "auto_target_error_trigger_rate",
    "auto_non_target_error_leakage",
    "auto_dominant_error_match_rate",
    "auto_answerability_rate",
    "auto_naturalness_mean",
    "auto_anchor_score_mean",
    "auto_anchor_difficulty_bucket",
    "auto_review_score",
    "auto_suggested_decision",
    "auto_reasons",
    "required_checks",
    "reviewer",
    "decision",
    "confirmed_target_error_type",
    "confirmed_scenario_type",
    "reference_answer_supported",
    "final_state_is_correctly_specified",
    "verifier_design_is_feasible",
    "reward_should_be_verifiable",
    "query_is_natural",
    "time_metadata_correct",
    "is_single_target_error",
    "release_priority",
    "issue_tags",
    "notes",
]

BOOL_COLUMNS = {
    "reference_answer_supported",
    "final_state_is_correctly_specified",
    "verifier_design_is_feasible",
    "reward_should_be_verifiable",
    "query_is_natural",
    "time_metadata_correct",
    "is_single_target_error",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert review tasks between JSONL and Excel")
    subparsers = parser.add_subparsers(dest="command", required=True)

    export_parser = subparsers.add_parser("export", help="Export JSONL review tasks to xlsx")
    export_parser.add_argument("--input", required=True, help="Input JSONL path")
    export_parser.add_argument("--output", required=True, help="Output xlsx path")

    import_parser = subparsers.add_parser("import", help="Import xlsx back to review JSONL")
    import_parser.add_argument("--input", required=True, help="Input xlsx path")
    import_parser.add_argument("--output", required=True, help="Output JSONL path")

    return parser.parse_args()


def _join_lines(value) -> str:
    if not value:
        return ""
    if isinstance(value, list):
        return "\n".join(str(x) for x in value)
    return str(value)


def _to_bool(value):
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    raise ValueError(f"无法解析布尔值: {value}")


def _build_row(task: Dict[str, object]) -> List[object]:
    auto = task.get("auto_signals", {})
    review = task.get("review_result", {})
    recommendation = task.get("auto_recommendation", {})
    return [
        task.get("item_id", ""),
        task.get("target_error_type", ""),
        task.get("scenario_type", ""),
        task.get("plan_id", ""),
        task.get("query", ""),
        task.get("context_preview", ""),
        task.get("reference_answer", ""),
        auto.get("target_error_trigger_rate", ""),
        auto.get("non_target_error_leakage", ""),
        auto.get("dominant_error_match_rate", ""),
        auto.get("answerability_rate", ""),
        auto.get("naturalness_mean", ""),
        auto.get("anchor_score_mean", ""),
        auto.get("anchor_difficulty_bucket", ""),
        recommendation.get("auto_review_score", ""),
        recommendation.get("suggested_decision", ""),
        _join_lines(recommendation.get("reasons", [])),
        _join_lines(task.get("required_checks", [])),
        review.get("reviewer", ""),
        review.get("decision", ""),
        review.get("confirmed_target_error_type", ""),
        review.get("confirmed_scenario_type", ""),
        review.get("reference_answer_supported", ""),
        review.get("final_state_is_correctly_specified", ""),
        review.get("verifier_design_is_feasible", ""),
        review.get("reward_should_be_verifiable", ""),
        review.get("query_is_natural", ""),
        review.get("time_metadata_correct", ""),
        review.get("is_single_target_error", ""),
        review.get("release_priority", ""),
        _join_lines(review.get("issue_tags", [])),
        review.get("notes", ""),
    ]


def export_to_excel(input_path: Path, output_path: Path) -> None:
    tasks = load_jsonl(input_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "review_tasks"
    ws.append(HEADERS)
    for task in tasks:
        ws.append(_build_row(task))

    ws.freeze_panes = "A2"
    widths = {
        "A": 16,
        "B": 14,
        "C": 12,
        "D": 28,
        "E": 32,
        "F": 48,
        "G": 28,
        "H": 10,
        "I": 10,
        "J": 10,
        "K": 10,
        "L": 10,
        "M": 10,
        "N": 12,
        "O": 10,
        "P": 12,
        "Q": 28,
        "R": 36,
        "S": 12,
        "T": 12,
        "U": 14,
        "V": 14,
        "W": 14,
        "X": 14,
        "Y": 14,
        "Z": 14,
        "AA": 14,
        "AB": 14,
        "AC": 14,
        "AD": 14,
        "AE": 20,
        "AF": 36,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def import_from_excel(input_path: Path, output_path: Path) -> None:
    wb = load_workbook(input_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    if header != HEADERS:
        raise ValueError("Excel 表头与预期不一致，请勿修改列名或顺序")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in row):
            continue
        payload = dict(zip(HEADERS, row))
        review_result = {
            "reviewer": payload["reviewer"] or "",
            "decision": payload["decision"] or "",
            "confirmed_target_error_type": payload["confirmed_target_error_type"] or payload["target_error_type"] or "",
            "confirmed_scenario_type": payload["confirmed_scenario_type"] or payload["scenario_type"] or "",
            "reference_answer_supported": _to_bool(payload["reference_answer_supported"]),
            "final_state_is_correctly_specified": _to_bool(payload["final_state_is_correctly_specified"]),
            "verifier_design_is_feasible": _to_bool(payload["verifier_design_is_feasible"]),
            "reward_should_be_verifiable": _to_bool(payload["reward_should_be_verifiable"]),
            "query_is_natural": _to_bool(payload["query_is_natural"]),
            "time_metadata_correct": _to_bool(payload["time_metadata_correct"]),
            "is_single_target_error": _to_bool(payload["is_single_target_error"]),
            "release_priority": payload["release_priority"] or "medium",
            "issue_tags": [x.strip() for x in str(payload["issue_tags"] or "").split("\n") if x.strip()],
            "notes": payload["notes"] or "",
        }
        result = {
            "item_id": payload["item_id"],
            "target_error_type": payload["target_error_type"],
            "scenario_type": payload["scenario_type"],
            "review_result": review_result,
        }
        rows.append(result)

    write_jsonl(output_path, rows)


def main() -> None:
    args = parse_args()
    if args.command == "export":
        export_to_excel(Path(args.input), Path(args.output))
        return
    import_from_excel(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
