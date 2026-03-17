"""Excel bridge for response/sentence annotation tasks."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook

from human_review.io_utils import load_jsonl, write_jsonl


HEADERS = [
    "annotation_id",
    "response_id",
    "query_id",
    "model_name",
    "intended_failure_mode",
    "query",
    "sentence_index",
    "sentence_text",
    "prefill_is_hallucinated",
    "prefill_attribution_type",
    "prefill_evidence_support",
    "annotator",
    "is_hallucinated",
    "attribution_type",
    "evidence_support",
    "severity",
    "notes",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert response/sentence annotation tasks between JSONL and Excel")
    subparsers = parser.add_subparsers(dest="command", required=True)

    export_parser = subparsers.add_parser("export")
    export_parser.add_argument("--input", required=True)
    export_parser.add_argument("--output", required=True)

    import_parser = subparsers.add_parser("import")
    import_parser.add_argument("--input", required=True)
    import_parser.add_argument("--output", required=True)

    return parser.parse_args()


def _to_bool(value):
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    raise ValueError(f"无法解析布尔值: {value}")


def export_to_excel(input_path: Path, output_path: Path) -> None:
    rows = load_jsonl(input_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "sentence_annotations"
    ws.append(HEADERS)
    for row in rows:
        result = row.get("annotation_result", {})
        ws.append(
            [
                row.get("annotation_id", ""),
                row.get("response_id", ""),
                row.get("query_id", ""),
                row.get("model_name", ""),
                row.get("intended_failure_mode", ""),
                row.get("query", ""),
                row.get("sentence_index", ""),
                row.get("sentence_text", ""),
                row.get("prefill_is_hallucinated", ""),
                row.get("prefill_attribution_type", ""),
                row.get("prefill_evidence_support", ""),
                result.get("annotator", ""),
                result.get("is_hallucinated", ""),
                result.get("attribution_type", ""),
                result.get("evidence_support", ""),
                result.get("severity", ""),
                result.get("notes", ""),
            ]
        )

    ws.freeze_panes = "A2"
    for col, width in {
        "A": 24,
        "B": 22,
        "C": 18,
        "D": 16,
        "E": 16,
        "F": 28,
        "G": 10,
        "H": 56,
        "I": 12,
        "J": 16,
        "K": 32,
        "L": 14,
        "M": 14,
        "N": 16,
        "O": 24,
        "P": 12,
        "Q": 28,
    }.items():
        ws.column_dimensions[col].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def import_from_excel(input_path: Path, output_path: Path) -> None:
    wb = load_workbook(input_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    if header != HEADERS:
        raise ValueError("Excel 表头与预期不一致，请勿修改列名或顺序")

    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if all(value in (None, "") for value in values):
            continue
        payload = dict(zip(HEADERS, values))
        rows.append(
            {
                "annotation_id": payload["annotation_id"],
                "response_id": payload["response_id"],
                "query_id": payload["query_id"],
                "model_name": payload["model_name"],
                "intended_failure_mode": payload["intended_failure_mode"],
                "query": payload["query"],
                "sentence_index": payload["sentence_index"],
                "sentence_text": payload["sentence_text"],
                "prefill_is_hallucinated": _to_bool(payload["prefill_is_hallucinated"]),
                "prefill_attribution_type": payload["prefill_attribution_type"] or "",
                "prefill_evidence_support": payload["prefill_evidence_support"] or "",
                "annotation_result": {
                    "annotator": payload["annotator"] or "",
                    "is_hallucinated": _to_bool(payload["is_hallucinated"]),
                    "attribution_type": payload["attribution_type"] or "",
                    "evidence_support": payload["evidence_support"] or "",
                    "severity": payload["severity"] or "",
                    "notes": payload["notes"] or "",
                },
            }
        )

    write_jsonl(output_path, rows)


def main() -> None:
    args = parse_args()
    if args.command == "export":
        export_to_excel(Path(args.input), Path(args.output))
        return
    import_from_excel(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
